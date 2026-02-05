import os
import logging
import pyexcel_ods
from .get_knowledge_base_config import get_knowledge_base_config

logger = logging.getLogger(__name__)

def get_section_name(knowledge_id):
    config = get_knowledge_base_config(knowledge_id)
    filepath = config.get('file_path')

    if not os.path.exists(filepath) or os.path.isdir(filepath):
        # logger.error(f"File '{filepath}' does not exist.")
        return knowledge_id
    
    if filepath.endswith('.md'):
        return knowledge_id

    if filepath.endswith('.xlsx'):
        from openpyxl import load_workbook
        try:
            wb = load_workbook(filepath, read_only=True)
            if wb.sheetnames:
                return wb.sheetnames[0]
        except Exception as e:
            logger.error(f"Error reading XLSX file '{filepath}': {e}")
            return knowledge_id

    try:
        book = pyexcel_ods.get_data(filepath)
    except Exception as e:
        logger.error(f"Error reading ODS file '{filepath}': {e}")
        return knowledge_id

    if not book:
        # logger.error(f"Sheet file '{filepath}' is empty or corrupted.")
        return knowledge_id
    
    section_name = list(book.keys())[0] # Use the first sheet if section_name is None
    # logger.info(f"No section_name provided, using the first sheet: '{section_name}'.")

    return section_name
