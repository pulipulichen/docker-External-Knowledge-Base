from __future__ import annotations
import base64
import json
import os
import sys
import logging
from collections import OrderedDict
from typing import Any, Dict, List, Optional

import zipfile
import io
import odf.opendocument
from odf.table import Table, TableRow, TableCell
from odf.draw import Frame, Image
from odf.text import P
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ....image_describe.image_describe import image_describe

logger = logging.getLogger(__name__)

def _get_headers(ws: Worksheet, header_row: int = 1) -> List[str]:
    """獲取指定列作為標頭，處理空白標頭並確保唯一性。"""
    headers: List[str] = []
    for cell in ws[header_row]:
        h = str(cell.value).strip() if cell.value is not None else ""
        headers.append(h)
    return headers

def _col_index_to_header(headers: List[str], col_1based: int) -> str:
    """根據列索引返回標頭名稱，若無標頭則回傳預設名稱。"""
    idx = col_1based - 1
    if 0 <= idx < len(headers) and headers[idx]:
        return headers[idx]
    return f"COL_{col_1based}"

def _collect_images(ws: Worksheet) -> Dict[tuple[int, int], List[str]]:
    """提取工作表中的圖片並轉換為 Base64。"""
    image_data: Dict[tuple[int, int], List[str]] = {}
    for img in getattr(ws, "_images", []):
        try:
            # 獲取圖片所在的行與列 (1-based)
            row = img.anchor._from.row + 1
            col = img.anchor._from.col + 1
            
            # 讀取圖片位元組並編碼
            raw_bytes = img._data()
            b64_str = base64.b64encode(raw_bytes).decode("utf-8")
            
            image_data.setdefault((row, col), []).append(b64_str)
        except Exception:
            continue
    return image_data

def _get_ods_cell_text(cell: TableCell) -> str:
    """提取 ODS 單元格中的文字內容。"""
    texts = []
    for p in cell.getElementsByType(P):
        texts.append(str(p))
    return "\n".join(texts).strip()

def _collect_ods_images(filepath: str) -> Dict[tuple[int, int], List[str]]:
    """提取 ODS 檔案中的圖片並轉換為 Base64。"""
    image_data: Dict[tuple[int, int], List[str]] = {}
    try:
        doc = odf.opendocument.load(filepath)
        zf = zipfile.ZipFile(filepath)
        
        # 遍歷所有工作表
        for sheet in doc.spreadsheet.getElementsByType(Table):
            rows = sheet.getElementsByType(TableRow)
            current_row = 1
            for row in rows:
                num_rows_repeated = int(row.getAttribute('numberrowsrepeated') or 1)
                cells = row.getElementsByType(TableCell)
                current_col = 1
                for cell in cells:
                    num_cols_repeated = int(cell.getAttribute('numbercolumnsrepeated') or 1)
                    
                    frames = cell.getElementsByType(Frame)
                    for frame in frames:
                        images = frame.getElementsByType(Image)
                        for img in images:
                            href = img.getAttribute('href')
                            if href and href.startswith('Pictures/'):
                                try:
                                    raw_bytes = zf.read(href)
                                    b64_str = base64.b64encode(raw_bytes).decode("utf-8")
                                    # 記錄到當前行與列
                                    print(f"Found image at ({current_row}, {current_col})")
                                    image_data.setdefault((current_row, current_col), []).append(b64_str)
                                except Exception:
                                    continue
                    current_col += num_cols_repeated
                current_row += num_rows_repeated
    except Exception as e:
        logger.error(f"Error collecting ODS images: {e}")
    return image_data

def _process_ods(filepath: str, section_name: Optional[str] = None, include_fields: Optional[List[str]] = None, header_row: int = 1) -> List[OrderedDict]:
    """處理 ODS 檔案，包含圖片描述。"""
    try:
        doc = odf.opendocument.load(filepath)
    except Exception as e:
        logger.error(f"Error loading ODS file '{filepath}': {e}")
        return []

    # 尋找目標工作表
    target_sheet = None
    all_sheets = doc.spreadsheet.getElementsByType(Table)
    if not all_sheets:
        logger.error(f"ODS file '{filepath}' has no sheets.")
        return []

    if section_name:
        for sheet in all_sheets:
            if sheet.getAttribute('name') == section_name:
                target_sheet = sheet
                break
        if not target_sheet:
            logger.error(f"Sheet '{section_name}' not found in ODS file.")
            return []
    else:
        target_sheet = all_sheets[0]

    # 獲取圖片
    images_map = _collect_ods_images(filepath)
    
    rows = target_sheet.getElementsByType(TableRow)
    headers: List[str] = []
    result_rows = []
    include_fields_list = include_fields or []
    
    current_row_idx = 0
    data_start_row = header_row
    
    # 遍歷行以獲取標頭和資料
    row_counter = 1
    for row in rows:
        num_rows_repeated = int(row.getAttribute('numberrowsrepeated') or 1)
        # 處理標頭
        if row_counter <= header_row < row_counter + num_rows_repeated:
            cells = row.getElementsByType(TableCell)
            current_col = 1
            for cell in cells:
                num_cols_repeated = int(cell.getAttribute('numbercolumnsrepeated') or 1)
                text = _get_ods_cell_text(cell)
                for _ in range(num_cols_repeated):
                    headers.append(text)
                current_col += num_cols_repeated
        
        # 處理資料行
        elif row_counter > header_row:
            for r_offset in range(num_rows_repeated):
                row_dict = OrderedDict()
                is_empty = True
                
                cells = row.getElementsByType(TableCell)
                current_col = 1
                for cell in cells:
                    num_cols_repeated = int(cell.getAttribute('numbercolumnsrepeated') or 1)
                    
                    key_idx = current_col - 1
                    key = headers[key_idx] if key_idx < len(headers) and headers[key_idx] else f"COL_{current_col}"
                    
                    if not include_fields_list or key in include_fields_list:
                        text_val = _get_ods_cell_text(cell)
                        if text_val:
                            is_empty = False
                        
                        # 處理圖片
                        imgs = images_map.get((row_counter + r_offset, current_col))
                        if imgs:
                            print(f"Processing {len(imgs)} images for row {row_counter + r_offset}, col {current_col}")
                        image_descriptions = []
                        if imgs:
                            for img_b64 in imgs:
                                try:
                                    desc = image_describe(img_b64)
                                    print(f"Image description: {desc[:50]}...")
                                    if desc and not desc.startswith("Error:"):
                                        image_descriptions.append(desc)
                                except Exception as e:
                                    print(f"Image describe exception: {e}")
                                    continue
                            if image_descriptions:
                                is_empty = False
                        
                        final_val = None
                        if text_val and image_descriptions:
                            desc_text = "\n".join(image_descriptions)
                            final_val = {"text": text_val, "images_description": desc_text}
                        elif text_val:
                            final_val = text_val
                        elif image_descriptions:
                            final_val = "\n".join(image_descriptions)
                        
                        if final_val is not None:
                            # 考慮列重複
                            for c_offset in range(num_cols_repeated):
                                current_key_idx = current_col + c_offset - 1
                                current_key = headers[current_key_idx] if current_key_idx < len(headers) and headers[current_key_idx] else f"COL_{current_col + c_offset}"
                                if not include_fields_list or current_key in include_fields_list:
                                    row_dict[current_key] = final_val

                    current_col += num_cols_repeated
                
                if not is_empty:
                    row_dict['__row_index__'] = row_counter + r_offset - header_row
                    result_rows.append(row_dict)

        row_counter += num_rows_repeated

    return result_rows

def _process_xlsx(filepath: str, section_name: Optional[str] = None, include_fields: Optional[List[str]] = None, header_row: int = 1) -> List[OrderedDict]:
    """處理 XLSX 檔案，包含圖片描述。"""
    wb = load_workbook(filepath, data_only=True)
    ws = wb[section_name] if section_name else wb.active

    headers = _get_headers(ws, header_row)
    images_map = _collect_images(ws)
    
    data_start_row = header_row + 1
    max_row = ws.max_row
    max_col = max(ws.max_column or 0, len(headers))

    result_rows = []
    include_fields_list = include_fields if include_fields else []

    for r in range(data_start_row, max_row + 1):
        row_dict = OrderedDict()
        is_empty = True

        for c in range(1, max_col + 1):
            key = _col_index_to_header(headers, c)
            if include_fields_list and key not in include_fields_list:
                continue

            val = ws.cell(row=r, column=c).value
            text_val = str(val).strip() if val is not None and val != "" else ""
            if text_val:
                is_empty = False
            
            imgs = images_map.get((r, c))
            image_descriptions = []
            if imgs:
                for img_b64 in imgs:
                    try:
                        desc = image_describe(img_b64)
                        if desc and not desc.startswith("Error:"):
                            image_descriptions.append(desc)
                    except Exception:
                        continue
                if image_descriptions:
                    is_empty = False

            final_val = None
            if text_val and image_descriptions:
                desc_text = "\n".join(image_descriptions)
                final_val = {"text": text_val, "images_description": desc_text}
            elif text_val:
                final_val = text_val
            elif image_descriptions:
                final_val = "\n".join(image_descriptions)
            
            if final_val is not None:
                row_dict[key] = final_val

        if not is_empty:
            row_dict['__row_index__'] = r - header_row
            result_rows.append(row_dict)
    return result_rows

def sheet_to_json(filepath: str, section_name: Optional[str] = None, include_fields: Optional[List[str]] = None, header_row: int = 1) -> List[OrderedDict]:
    """
    Reads a sheet file (ODS or XLSX) and converts it into a list of dictionaries.
    
    Args:
        filepath: The path to the file.
        section_name: The name of the sheet to extract.
        include_fields: List of field names to include.
        header_row: The row index (1-based) where headers are located (for XLSX).
        
    Returns:
        A list of OrderedDict objects representing the rows.
    """
    if not os.path.exists(filepath):
        logger.error(f"File '{filepath}' does not exist.")
        return []
    
    ext = os.path.splitext(filepath)[1].lower()
    
    if ext == '.ods':
        return _process_ods(filepath, section_name, include_fields, header_row)
    elif ext == '.xlsx':
        return _process_xlsx(filepath, section_name, include_fields, header_row)
    else:
        logger.error(f"Unsupported file extension: {ext}")
        return []

