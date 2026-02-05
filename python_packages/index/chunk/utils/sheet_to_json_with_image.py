from __future__ import annotations

import base64
import json
import os
import sys
from typing import Any, Dict, List, Optional

# Add python_packages to sys.path to allow importing image_describe
current_dir = os.path.dirname(os.path.abspath(__file__))
# Navigate up to the python_packages directory
# path is: .../python_packages/index/chunk/utils/sheet_to_json_with_image.py
# target is: .../python_packages/
package_root = os.path.abspath(os.path.join(current_dir, "../../../../"))
if package_root not in sys.path:
    sys.path.append(package_root)

from image_describe.image_describe import image_describe
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _get_headers(ws: Worksheet, header_row: int = 1) -> List[str]:
    """獲取指定列作為標頭，處理空白標頭並確保唯一性。"""
    headers: List[str] = []
    for cell in ws[header_row]:
        h = str(cell.value).strip() if cell.value is not None else ""
        headers.append(h)
    return headers


def _col_index_to_header(headers: List[str], col_1based: int) -> str:
    """根據列索引返回標頭名稱，若無標頭則回傳預設名稱。"""
    idx = col_1based - 1
    if 0 <= idx < len(headers) and headers[idx]:
        return headers[idx]
    return f"COL_{col_1based}"


def _collect_images(ws: Worksheet) -> Dict[tuple[int, int], List[str]]:
    """提取工作表中的圖片並轉換為 Base64。"""
    image_data: Dict[tuple[int, int], List[str]] = {}
    for img in getattr(ws, "_images", []):
        try:
            # 獲取圖片所在的行與列 (1-based)
            row = img.anchor._from.row + 1
            col = img.anchor._from.col + 1
            
            # 讀取圖片位元組並編碼
            raw_bytes = img._data()
            b64_str = base64.b64encode(raw_bytes).decode("utf-8")
            
            image_data.setdefault((row, col), []).append(b64_str)
        except Exception:
            continue
    return image_data


def convert_xlsx_to_json(
    xlsx_path: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    indent: Optional[int] = None
) -> str:
    """
    主要進入點：將 XLSX 轉換為 JSON 字串。
    
    Args:
        xlsx_path: Excel 檔案路徑。
        sheet_name: 指定工作表名稱，預設為首個工作表。
        header_row: 標頭所在的行數（從 1 開始）。
        indent: JSON 輸出縮排格數，設為 None 則不排版（壓縮格式）。
        
    Returns:
        JSON 格式的字串。
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = _get_headers(ws, header_row)
    images_map = _collect_images(ws)
    
    data_start_row = header_row + 1
    max_row = ws.max_row
    max_col = max(ws.max_column or 0, len(headers))

    result_rows = []

    for r in range(data_start_row, max_row + 1):
        row_item: Dict[str, Any] = {}
        is_empty = True

        for c in range(1, max_col + 1):
            key = _col_index_to_header(headers, c)
            val = ws.cell(row=r, column=c).value
            
            # 處理文字資料
            text_val = str(val).strip() if val is not None and val != "" else ""
            if text_val:
                is_empty = False
            
            # 處理圖片資料
            imgs = images_map.get((r, c))
            image_descriptions = []
            if imgs:
                for img_b64 in imgs:
                    try:
                        desc = image_describe(img_b64)
                        # 如果描述成功且不是錯誤訊息，則加入
                        if desc and not desc.startswith("Error:"):
                            image_descriptions.append(desc)
                    except Exception:
                        continue
                
                if image_descriptions:
                    is_empty = False

            # 組合最終值
            final_val = None
            if text_val and image_descriptions:
                desc_text = "\n".join(image_descriptions)
                final_val = {"text": text_val, "images_description": desc_text}
            elif text_val:
                final_val = text_val
            elif image_descriptions:
                final_val = "\n".join(image_descriptions)
            
            if final_val is not None:
                row_item[key] = final_val

        if not is_empty:
            row_item['__row_index__'] = r - header_row + 1
            result_rows.append(row_item)

    # 序列化為 JSON，預設不使用縮排 (indent=None)
    return json.dumps(result_rows, ensure_ascii=False, indent=indent, default=str)


# --- 測試範例 ---
if __name__ == "__main__":
    # 使用範例
    try:
        # 現在呼叫此函式，預設會輸出無縮排的單行 JSON
        # json_output = convert_xlsx_to_json("test.xlsx")
        print("函式已就緒，預設不排版。若需排版請傳入 indent=2。")
    except Exception as e:
        print(f"執行錯誤: {e}")