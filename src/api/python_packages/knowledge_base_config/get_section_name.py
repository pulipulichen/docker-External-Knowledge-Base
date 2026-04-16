import os
import logging
import zipfile
import xml.etree.ElementTree as ET

import pyexcel_ods
from .get_knowledge_base_config import get_knowledge_base_config

logger = logging.getLogger(__name__)

# OOXML workbook namespace (sheet names live in xl/workbook.xml; no need for full openpyxl parse).
_XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


from openpyxl import load_workbook

def _first_sheet_name_from_xlsx(file_path):
    

    # 如果 filepath 是連接檔，那就取得原始檔案路徑後再來輸入
    if os.path.islink(file_path):
        # Resolve symlink to actual file path
        file_path = os.path.realpath(file_path)

    os.system(f"cat '{file_path}' > /dev/null")
    os.system(f"cp '{file_path}' /tmp")
    file_path = os.path.join('/tmp', os.path.basename(file_path))
    
    logger.info(f"Reading XLSX file '{file_path}' in _first_sheet_name_from_xlsx")
    wb = load_workbook(file_path, read_only=True)
    logger.info(f"Sheetnames: '{wb.sheetnames}'")

    return wb.sheetnames[0]


def get_section_name(knowledge_id):
    config = get_knowledge_base_config(knowledge_id)
    filepath = config.get('file_path')

    logger.info(f"Filepath: '{filepath}'")

    if filepath is None or not os.path.exists(filepath) or os.path.isdir(filepath):
        # logger.error(f"File '{filepath}' does not exist.")
        return knowledge_id
    
    if filepath.endswith('.md'):
        return knowledge_id

    if filepath.endswith('.xlsx'):
        try:
            logger.info(f"Reading XLSX file '{filepath}'")
            first = _first_sheet_name_from_xlsx(filepath)
            logger.info(f"First sheet name: '{first}'")
            return first if first is not None else knowledge_id
        except Exception as e:
            logger.error(f"Error reading XLSX file '{filepath}': {e}")
            return knowledge_id

    try:
        book = pyexcel_ods.get_data(filepath)
    except Exception as e:
        logger.error(f"Error reading ODS file '{filepath}': {e}")
        return knowledge_id

    if not book:
        # logger.error(f"Sheet file '{filepath}' is empty or corrupted.")
        return knowledge_id
    
    section_name = list(book.keys())[0] # Use the first sheet if section_name is None
    # logger.info(f"No section_name provided, using the first sheet: '{section_name}'.")

    return section_name
